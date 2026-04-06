from abc import ABC, abstractmethod
from pathlib import Path
from typing import Dict, Any
import openpyxl


class BaseExcelLoader(ABC):
    """
    Базовый класс для загрузчиков Excel.
    Каждая задача реализует свой загрузчик.
    """

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        self._workbook = None
        self._cache = {}
        self._load_workbook()

    def _load_workbook(self):
        """Загружает workbook при создании"""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Файл не найден: {self.excel_path}")
        self._workbook = openpyxl.load_workbook(self.excel_path, data_only=True)

    def _get_sheet_data(self, sheet_name: str) -> Dict[int, Dict]:
        """
        Загружает данные с листа в словарь {digit: row_data}
        """
        if sheet_name in self._cache:
            return self._cache[sheet_name]

        if sheet_name not in self._workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в файле {self.excel_path}")

        sheet = self._workbook[sheet_name]
        data = {}

        # Читаем заголовки из первой строки
        headers = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(cell).lower().strip() if cell else f"col_{i}" for i, cell in enumerate(row)]
                continue

            if row[0] is None:
                continue

            digit = int(row[0])
            row_data = {}
            for i, value in enumerate(row[1:], 1):
                if i < len(headers):
                    key = headers[i]
                else:
                    key = f"col_{i}"

                # Преобразуем в число, если возможно
                if isinstance(value, (int, float)):
                    row_data[key] = float(value) if '.' in str(value) else int(value)
                elif value is not None:
                    row_data[key] = value

            data[digit] = row_data

        self._cache[sheet_name] = data
        return data

    @abstractmethod
    def load_cipher(self, cipher: str) -> Dict[str, Any]:
        """
        Загружает параметры по шифру.
        Каждая задача реализует свою логику.
        """
        pass

    def print_summary(self):
        """Выводит информацию о файле"""
        print(f"\n📊 Excel файл: {self.excel_path.name}")
        print(f"   Листы: {self._workbook.sheetnames}")